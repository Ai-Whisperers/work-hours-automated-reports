"""Report generator implementations."""

from pathlib import Path
from typing import Dict, Any, Optional
import logging
from datetime import datetime

import polars as pl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from jinja2 import Template

from ...application.ports import ReportGenerator


logger = logging.getLogger(__name__)


class ExcelReportGenerator(ReportGenerator):
    """Excel report generator implementation."""
    
    async def generate(
        self,
        data: Dict[str, Any],
        format: str,
        output_path: Optional[Path] = None,
        options: Optional[Dict[str, Any]] = None
    ) -> Path:
        """Generate an Excel report.
        
        Args:
            data: Report data
            format: Output format (should be 'excel')
            output_path: Output file path
            options: Additional options
            
        Returns:
            Path to generated report
        """
        if format != "excel":
            raise ValueError(f"ExcelReportGenerator only supports 'excel' format, got '{format}'")
        
        # Default output path
        if not output_path:
            output_path = Path(f"report_{datetime.now():%Y%m%d_%H%M%S}.xlsx")
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheets
        self._create_summary_sheet(wb, data, options)
        self._create_by_person_sheet(wb, data)
        self._create_by_work_item_sheet(wb, data)
        self._create_raw_data_sheet(wb, data)
        
        # Save workbook
        wb.save(output_path)
        
        logger.info(f"Excel report generated: {output_path}")
        return output_path
    
    def _create_summary_sheet(self, wb: Workbook, data: Dict[str, Any], options: Optional[Dict[str, Any]]):
        """Create summary sheet."""
        ws = wb.create_sheet("Summary")

        # Title
        ws["A1"] = "Clockify-ADO Time Tracking Report"
        ws["A1"].font = Font(bold=True, size=14)
        
        # Date range
        if options and "date_range" in options:
            ws["A3"] = "Report Period:"
            ws["B3"] = options["date_range"]
        
        # Statistics
        row = 5
        ws[f"A{row}"] = "Statistics"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        
        row += 1
        stats = [
            ("Total Entries", data.get("total_entries", 0)),
            ("Matched Entries", data.get("match_count", 0)),
            ("Unmatched Entries", data.get("unmatch_count", 0)),
            ("Match Rate", f"{(data.get('match_count', 0) / max(data.get('total_entries', 1), 1)) * 100:.1f}%"),
        ]
        
        for label, value in stats:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = value
            row += 1
        
        # Adjust column widths
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 15
    
    def _create_by_person_sheet(self, wb: Workbook, data: Dict[str, Any]):
        """Create sheet grouped by person."""
        ws = wb.create_sheet("ByPerson")
        
        # Convert data to DataFrame
        entries = data.get("matched_entries", [])
        if not entries:
            ws["A1"] = "No matched entries found"
            return
        
        df = pl.DataFrame(entries)
        
        # Group by person and work item
        grouped = df.group_by(["user_name", "work_item_id", "work_item_title"]).agg([
            pl.col("duration_hours").sum().alias("total_hours"),
            pl.col("id").count().alias("entry_count")
        ]).sort(["user_name", "total_hours"], descending=[False, True])
        
        # Write headers
        headers = ["User", "Work Item ID", "Work Item Title", "Total Hours", "Entry Count"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Write data
        for row_idx, row in enumerate(grouped.iter_rows(), 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column].width = min(max_length + 2, 50)
    
    def _create_by_work_item_sheet(self, wb: Workbook, data: Dict[str, Any]):
        """Create sheet grouped by work item."""
        ws = wb.create_sheet("ByWorkItem")
        
        entries = data.get("matched_entries", [])
        if not entries:
            ws["A1"] = "No matched entries found"
            return
        
        df = pl.DataFrame(entries)
        
        # Group by work item
        grouped = df.group_by(["work_item_id", "work_item_title", "work_item_type"]).agg([
            pl.col("duration_hours").sum().alias("total_hours"),
            pl.col("user_name").n_unique().alias("unique_users"),
            pl.col("id").count().alias("entry_count")
        ]).sort("total_hours", descending=True)
        
        # Write headers
        headers = ["Work Item ID", "Title", "Type", "Total Hours", "Unique Users", "Entry Count"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Write data
        for row_idx, row in enumerate(grouped.iter_rows(), 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column].width = min(max_length + 2, 50)
    
    def _create_raw_data_sheet(self, wb: Workbook, data: Dict[str, Any]):
        """Create raw data sheet."""
        ws = wb.create_sheet("RawData")
        
        all_entries = data.get("matched_entries", []) + data.get("unmatched_entries", [])
        if not all_entries:
            ws["A1"] = "No entries found"
            return
        
        # Get all unique keys
        all_keys = set()
        for entry in all_entries:
            all_keys.update(entry.keys())
        
        # Sort keys for consistent ordering
        headers = sorted(all_keys)
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Write data
        for row_idx, entry in enumerate(all_entries, 2):
            for col_idx, header in enumerate(headers, 1):
                value = entry.get(header)
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column].width = min(max_length + 2, 50)
    
    def supports_format(self, format: str) -> bool:
        """Check if format is supported."""
        return format == "excel"
    
    async def validate_data(self, data: Dict[str, Any]) -> bool:
        """Validate report data."""
        required_keys = ["matched_entries", "unmatched_entries", "total_entries"]
        return all(key in data for key in required_keys)


class HTMLReportGenerator(ReportGenerator):
    """HTML report generator implementation."""
    
    async def generate(
        self,
        data: Dict[str, Any],
        format: str,
        output_path: Optional[Path] = None,
        options: Optional[Dict[str, Any]] = None
    ) -> Path:
        """Generate an HTML report.
        
        Args:
            data: Report data
            format: Output format (should be 'html')
            output_path: Output file path
            options: Additional options
            
        Returns:
            Path to generated report
        """
        if format != "html":
            raise ValueError(f"HTMLReportGenerator only supports 'html' format, got '{format}'")
        
        # Default output path
        if not output_path:
            output_path = Path(f"report_{datetime.now():%Y%m%d_%H%M%S}.html")
        
        # Create HTML template
        template = Template(self._get_html_template())
        
        # Prepare template data
        template_data = self._prepare_template_data(data, options)
        
        # Render HTML
        html_content = template.render(**template_data)
        
        # Save to file
        output_path.write_text(html_content)
        
        logger.info(f"HTML report generated: {output_path}")
        return output_path
    
    def _get_html_template(self) -> str:
        """Get HTML template."""
        return '''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Time Tracking Report</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body {
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Oxygen, Ubuntu, sans-serif;
            background: #1a1a2e;
            color: #eee;
            padding: 2rem;
        }
        .container { max-width: 1200px; margin: 0 auto; }
        h1 { color: #4fbdba; margin-bottom: 2rem; }
        h2 { color: #7ec8e3; margin: 2rem 0 1rem; }
        .stats {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 1rem;
            margin-bottom: 2rem;
        }
        .stat-card {
            background: #16213e;
            padding: 1.5rem;
            border-radius: 8px;
            border-left: 4px solid #4fbdba;
        }
        .stat-value { font-size: 2rem; font-weight: bold; color: #7ec8e3; }
        .stat-label { color: #888; margin-top: 0.5rem; }
        table {
            width: 100%;
            border-collapse: collapse;
            background: #16213e;
            border-radius: 8px;
            overflow: hidden;
            margin-bottom: 2rem;
        }
        th {
            background: #0f3460;
            color: #4fbdba;
            padding: 1rem;
            text-align: left;
        }
        td { padding: 1rem; border-top: 1px solid #0f3460; }
        tr:hover { background: #1e3a5f; }
        .date-range { color: #7ec8e3; margin-bottom: 1rem; }
    </style>
</head>
<body>
    <div class="container">
        <h1>Time Tracking Report</h1>
        <div class="date-range">{{ date_range }}</div>
        
        <div class="stats">
            <div class="stat-card">
                <div class="stat-value">{{ total_entries }}</div>
                <div class="stat-label">Total Entries</div>
            </div>
            <div class="stat-card">
                <div class="stat-value">{{ match_rate }}%</div>
                <div class="stat-label">Match Rate</div>
            </div>
            <div class="stat-card">
                <div class="stat-value">{{ total_hours }}h</div>
                <div class="stat-label">Total Hours</div>
            </div>
        </div>
        
        <h2>Top Work Items</h2>
        <table>
            <thead>
                <tr>
                    <th>Work Item</th>
                    <th>Type</th>
                    <th>Hours</th>
                    <th>Contributors</th>
                </tr>
            </thead>
            <tbody>
                {% for item in top_work_items %}
                <tr>
                    <td>#{{ item.id }} - {{ item.title }}</td>
                    <td>{{ item.type }}</td>
                    <td>{{ item.hours }}</td>
                    <td>{{ item.contributors }}</td>
                </tr>
                {% endfor %}
            </tbody>
        </table>
        
        <h2>Top Contributors</h2>
        <table>
            <thead>
                <tr>
                    <th>User</th>
                    <th>Total Hours</th>
                    <th>Work Items</th>
                </tr>
            </thead>
            <tbody>
                {% for user in top_users %}
                <tr>
                    <td>{{ user.name }}</td>
                    <td>{{ user.hours }}</td>
                    <td>{{ user.work_items }}</td>
                </tr>
                {% endfor %}
            </tbody>
        </table>
    </div>
</body>
</html>'''
    
    def _prepare_template_data(self, data: Dict[str, Any], options: Optional[Dict[str, Any]]) -> Dict[str, Any]:
        """Prepare data for template."""
        entries = data.get("matched_entries", [])
        
        # Calculate statistics
        total_hours = sum(e.get("duration_hours", 0) for e in entries)
        
        # Group data for display
        work_items = {}
        users = {}
        
        for entry in entries:
            # Aggregate by work item
            wi_id = entry.get("work_item_id")
            if wi_id:
                if wi_id not in work_items:
                    work_items[wi_id] = {
                        "id": wi_id,
                        "title": entry.get("work_item_title", "Unknown"),
                        "type": entry.get("work_item_type", "Unknown"),
                        "hours": 0,
                        "contributors": set()
                    }
                work_items[wi_id]["hours"] += entry.get("duration_hours", 0)
                work_items[wi_id]["contributors"].add(entry.get("user_name"))
            
            # Aggregate by user
            user = entry.get("user_name")
            if user:
                if user not in users:
                    users[user] = {
                        "name": user,
                        "hours": 0,
                        "work_items": set()
                    }
                users[user]["hours"] += entry.get("duration_hours", 0)
                if wi_id:
                    users[user]["work_items"].add(wi_id)
        
        # Sort and limit
        top_work_items = sorted(work_items.values(), key=lambda x: x["hours"], reverse=True)[:10]
        top_users = sorted(users.values(), key=lambda x: x["hours"], reverse=True)[:10]
        
        # Format for template
        for item in top_work_items:
            item["hours"] = f"{item['hours']:.1f}"
            item["contributors"] = len(item["contributors"])
        
        for user in top_users:
            user["hours"] = f"{user['hours']:.1f}"
            user["work_items"] = len(user["work_items"])
        
        return {
            "date_range": options.get("date_range", "Unknown period") if options else "Unknown period",
            "total_entries": data.get("total_entries", 0),
            "match_rate": round((data.get("match_count", 0) / max(data.get("total_entries", 1), 1)) * 100, 1),
            "total_hours": round(total_hours, 1),
            "top_work_items": top_work_items,
            "top_users": top_users
        }
    
    def supports_format(self, format: str) -> bool:
        """Check if format is supported."""
        return format == "html"
    
    async def validate_data(self, data: Dict[str, Any]) -> bool:
        """Validate report data."""
        required_keys = ["matched_entries", "total_entries"]
        return all(key in data for key in required_keys)